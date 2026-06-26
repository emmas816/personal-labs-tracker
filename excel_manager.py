import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Design Theme
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark steel blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

PANEL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Soft light blue
PANEL_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")

STANDARD_FONT = Font(name="Calibri", size=11, bold=False, color="000000")
BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="000000")

THIN_BORDER_SIDE = Side(border_style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)

# Panels & Tests database
COMMON_PANELS = [
    {
        "name": "Complete Blood Count Panel with Differential",
        "tests": [
            ("WBC", "4.0-10.5 x 10³/µL", "5 - 7"),
            ("RBC", "men: 4.2-6.1\nwomen: 3.89-5.6 x 10³/µL", "within range"),
            ("Hemoglobin", "men: 14.0-17.5\nwomen: 12.3-15.3 g/dL", "upper end"),
            ("Hematocrit", "men: 41.5-50.4%\nwomen: 36.9-44.6%", "upper end"),
            ("MCV", "80-98 fL", "within range"),
            ("MCH", "27.0-34.0 pg", "within range"),
            ("MCHC", "32.0-36.0 g/dL", "within range"),
            ("RDW", "11.7-15.0%", "within range"),
            ("Platelets", "140-415 x 10³/µL", "175-250"),
            ("Neutrophils", "% not estab.", ""),
            ("Neutrophils (Absolute)", "1.8-7.8 x 10³/µL", "3.0-6.5"),
            ("Lymphocytes", "% not estab.", ""),
            ("Lymphs (Absolute)", "0.7-4.5 x 10³/µL", "2.5-4.5"),
            ("Neutrophils/Lymphocytes Ratio", "1 - 2", "~2 (example 50/25)"),
            ("Monocytes", "% not estab.", ""),
            ("Monocytes (Absolute)", "0.1-1.0 x 10³/µL", "within range"),
            ("Eosinophils", "% not estab.", ""),
            ("Eos (Absolute)", "0.0-0.4 x 10³/µL", "0.0-0.3"),
            ("Basophils", "% not estab.", ""),
            ("Baso (Absolute)", "0.0-0.1 x 10³/µL", "within range"),
            ("Immature Granulocytes", "0.0-0.1 x 10³/µL", "within range")
        ]
    },
    {
        "name": "Comprehensive Chemistry Panel",
        "tests": [
            ("Glucose", "65-99 mg/dL", "70-85"),
            ("Uric Acid", "2.5-7.1 mg/dL", "<6"),
            ("BUN", "5-26 mg/dL", "9 - 23"),
            ("Creatinine", "0.57-1.00 mg/dL", "within range"),
            ("eGFR (kidneys filtration rate)", "Greater than 59", "greater than 60"),
            ("BUN/Creatinine Ratio", "8-27", "15-27"),
            ("Sodium", "135-145 mmol/L", "140 - 142"),
            ("Potassium", "3.5-5.2 mmol/L", "~4.2"),
            ("Chloride", "97-108 mmol/L", "within range"),
            ("Carbon Dioxide (bicarbonate)", "20-32 mmol/L", "23 - 28"),
            ("Calcium", "8.7-10.2 mg/dL", "9.0-10.2"),
            ("Phosphorus", "2.5-4.5 mg/dL", "within range"),
            ("Total Protein", "6.0-8.5 g/dL", "7.0-8.5 (under 7 is metabolic wasting, sarcopenia)"),
            ("Albumin", "3.5-5.5 g/dL", "4.5-5.5 (+ under 4 is metabolic wasting, sarcopenia)"),
            ("Globulin", "1.5-4.5 g/dL", "within range"),
            ("A/G Ratio", "1.1-2.5", "within range"),
            ("Bilirubin", "0.0-1.2 mg/dL", "within range"),
            ("Alkaline Phosphatase (ALP)", "25-150 IU/L", "65-85"),
            ("LDH", "119-226 IU/L", "< 175 (if more than 175 mitochondria is not functioning)"),
            ("AST", "0-40 IU/L", "8 - 18 (if more than 20 sign of bone loss)"),
            ("ALT", "0-40 IU/L", "7 - 19"),
            ("Iron", "27-159 ug/dL", "40 - 100"),
            ("GGT", "0 - 60 IU/L", "men <20, women <10")
        ]
    },
    {
        "name": "Healthy Lipid Panel",
        "tests": [
            ("Total Cholesterol (TC)", "130-200 mg/dL", "145-165"),
            ("Triglycerides (TG)", "30-150 mg/dL", "< 50"),
            ("HDL", "35-150 mg/dL", "> 65"),
            ("VLDL", "5-40 mg/dL", "within range"),
            ("LDL", "0-129 mg/dL", "< 70"),
            ("TC/HDL", "0-4.4", "< 2.5"),
            ("TG/HDL", "0-3.7", "< 2.0 (~1)"),
            ("LDL/HDL", "0.5-3.0", "< 2.0 (~1)")
        ]
    },
    {
        "name": "Inflammation Marker Panel",
        "tests": [
            ("Sedimentation Rate", "0-30 mm/hr", "0-10"),
            ("Cardiac CRP", "0.00-3.00 mg/L", "< 1.0 women, < 0.55 men"),
            ("Homocysteine", "0.0-15.0 μmol/L", "0.0-6.0"),
            ("H-pylori", "negative 0.0-0.79", "negative")
        ]
    },
    {
        "name": "Healthy Thyroid Panel",
        "tests": [
            ("TSH", "0.4-5.5 mU/L", "1.0-2.0"),
            ("T4, total", "4.5-12.5 mcg/dL", "7.5-8.1"),
            ("T4, free", "0.8-1.8 ng/dL", "1.3 - 1.8"),
            ("T3, total", "60-181 ng/dL", "120-124"),
            ("FT3(free)", "2.60-4.80 pg/mL", "3.2 - 4.2"),
            ("T3, uptake", "24-39 %", "within range"),
            ("RT3(reverse)", "9.2 - 24.1 ng/dL", "9.2 - 21"),
            ("RT3/FT3", "< 5", "within range")
        ]
    },
    {
        "name": "Nutrient Panel",
        "tests": [
            ("Folic Acid (Folate)", "2.2-3.0 ng/mL", "> 3.0"),
            ("Iron", "27-159 ug/dL", "40 - 100"),
            ("Vitamin B12", "211-946 pg/mL", "600-1000"),
            ("Vitamin C", "0.4-2.0 mg/dL", "1.5-3.0"),
            ("Vitamin D, 25-Hydroxy", "32-100 ng/mL", "50 - 80"),
            ("Ferritin", "30-400 ng/mL", "40 - 90"),
            ("TIBC Transferrin binding capacity", "250-450 mcg/dL", "~350"),
            ("TSAT Transferrin Saturation", "25 - 35%", "<30%")
        ]
    },
    {
        "name": "Healthy Glucose Metabolic Panel",
        "tests": [
            ("Glucose", "65-99 mg/dL", "70 - 85"),
            ("Hemoglobin A1c", "4.8-5.6%", "< 5%"),
            ("Insulin", "0.0-24.9 µIU/mL", "2 - 5")
        ]
    }
]

MALE_HORMONE_PANEL = {
    "name": "Male Hormone Panel",
    "tests": [
        ("Estradiol (E2)", "21-54 pg/mL", "20-30"),
        ("Progesterone", "<0.5 ng/mL", "<0.5"),
        ("Testosterone", "264-916 ng/dL", "500 - 1000"),
        ("Free Testosterone", "4.7-26.5 pg/mL", "6.5 - 25"),
        ("DHT", "30-85 ng/dL", "30-50"),
        ("DHEA-S", "30-620 mcg/dL", "350-500"),
        ("Cortisol", "AM: 6.2-19.4 mcg/dL\nPM: 2.3-11.9 mcg/dL", "5-10"),
        ("PSA", "0.0-4.0 ng/mL", "0.0 - 1.0"),
        ("Sex Hormone Binding Globulin (SHBG)", "20-49 yrs: 16.5-55.9 nmol/L\n>49 yrs: 19.3-76.4 nmol/L", "30-40")
    ]
}

FEMALE_HORMONE_PANEL = {
    "name": "Female Hormone Panel",
    "tests": [
        ("Estradiol (E2)", 
         "12-498 pg/mL (age-matched range)\nPostmenopausal Women:\n<6.0-54.7 pg/mL\nLowest levels shown to ameliorate symptoms: 30-50 pg/mL", 
         "Pre-Menopausal Women:\n◦ Follicular: 12.5-166.0 pg/mL\n◦ Ovulation: 85.8-498 pg/mL\n◦ Luteal: 43.8-211.0 pg/mL\nPostmenopausal Women:\nWith typical Bi-est: 80-100 pg/mL"),
        ("Progesterone", 
         "0.1-27.0 ng/mL (age-matched range)", 
         "Pre-Menopausal Women:\n◦ Follicular: 0.1-0.9 ng/mL\n◦ Luteal: 1.8-23.9 ng/mL\n◦ Ovulatory: 0.1-12.0 ng/mL\n15-23 ng/mL at ~day 21\nPostmenopausal Women:\n2-6 ng/mL"),
        ("LH", "1.0-95.6 mIU/mL (age-matched range)", "within range"),
        ("FSH", "1.7-134.8 mIU/mL (age-matched range)", "25-100"),
        ("Testosterone", "6-82 ng/dL", "35-65"),
        ("DHEA-S", "30-560 mcg/dL (age-matched range)", "275-400"),
        ("Cortisol", "AM: 6.2-19.4 mcg/dL\nPM: 2.3-11.9 mcg/dL", "5-10")
    ]
}

def clean_name(name):
    """
    Normalizes test names by converting to lowercase, replacing punctuation with spaces,
    collapsing multiple spaces, and removing common medical units/suffixes.
    """
    if not name:
        return ""
    n = name.lower().strip()
    # Replace separators with spaces to avoid squishing words
    for char in [",", "/", "%", "(", ")", "-", "_", "[", "]"]:
        n = n.replace(char, " ")
    # Collapse multiple spaces
    n = " ".join(n.split())
    
    # Remove common units/suffixes
    suffixes = [
        "g/dl", "mg/dl", "mmol/l", "pg/ml", "ng/dl", "mcg/dl", 
        "ui/ml", "miu/ml", "μmol", "uieu/ml", "uie/ml", "ui/l",
        "uiu/ml", "miu/ml"
    ]
    for suffix in suffixes:
        if n.endswith(suffix):
            n = n[:-len(suffix)].strip()
    return n

def match_test_name(standard_name, parsed_name):
    """
    Fuzzy checks if the standard test name maps to the parsed name.
    """
    sc = clean_name(standard_name)
    pc = clean_name(parsed_name)
    
    if sc == pc:
        return True

    # 1. Differential absolute vs percentage protection rule:
    # A differential marker (Neutrophils, Lymphocytes, Monocytes, Eosinophils, Basophils)
    # in percentage form must not match absolute form.
    diff_markers = ["neutrophil", "lymphocyte", "monocyte", "eosinophil", "basophil", "lymph", "eos", "baso"]
    is_standard_diff = any(dm in sc for dm in diff_markers)
    is_parsed_diff = any(dm in pc for dm in diff_markers)
    if is_standard_diff and is_parsed_diff:
        is_standard_abs = ("absolute" in sc or "abs" in sc)
        is_parsed_abs = ("absolute" in pc or "abs" in pc)
        if is_standard_abs != is_parsed_abs:
            return False

    # 2. Ratio protection rule: a ratio test should only match another ratio test
    # We check if 'ratio' is a whole word in the cleaned name or if '/' was in the original name.
    is_standard_ratio = ("ratio" in sc.split() or "/" in standard_name)
    is_parsed_ratio = ("ratio" in pc.split() or "/" in parsed_name)
    if is_standard_ratio != is_parsed_ratio:
        return False
        
    # Map synonyms/aliases
    mappings = {
        "wbc": ["white blood cell", "white blood cells", "wbc count", "leukocytes"],
        "rbc": ["red blood cell", "red blood cells", "rbc count", "erythrocytes"],
        "hemoglobin": ["hgb", "hemoglobin g/dl"],
        "hematocrit": ["hct", "hematocrit %"],
        "platelets": ["platelet count", "plt"],
        "tsh": ["thyroid stimulating hormone"],
        "t3 total": ["triiodothyronine", "t3", "total t3"],
        "t4 total": ["thyroxine", "t4", "total t4"],
        "ft3 free": ["free t3", "ft3 free", "free triiodothyronine", "ft3"],
        "ft4 free": ["free t4", "ft4 free", "free thyroxine", "ft4", "t4 free"],
        "estradiol e2": ["estradiol", "e2", "estradiol, sensitive"],
        "progesterone": ["prog", "progesterone, LC/MS"],
        "testosterone": ["testo", "total testosterone", "testosterone, total", "testosterone total"],
        "free testosterone": ["free testo", "testosterone, free", "testosterone free"],
        "dhea s": ["dhea sulfate", "dheas", "dhea-sulfate", "dhea, sulfate"],
        "shbg": ["sex hormone binding globulin", "shbg, serum"],
        "egfr kidneys filtration rate": [
            "egfr", "e-gfr", "glomerular filtration rate", "estimated gfr", 
            "egfr non-african american", "egfr african american", "egfr (kidney)", "egfr calc"
        ],
        "bun creatinine ratio": ["bun/creatinine ratio", "bun/creat ratio", "bun/creatinine", "bun/creatine ratio", "buncreatine ratio"],
        "carbon dioxide bicarbonate": [
            "carbon dioxide", "bicarbonate", "co2", "carbon dioxide total", 
            "carbon dioxide, total", "co2 total", "carbon dioxide (bicarbonate)"
        ],
        "alkaline phosphatase alp": ["alkaline phosphatase", "alk phos", "alp"],
        "folic acid folate": ["folate", "folic acid", "folate serum"],
        "vitamin d 25 hydroxy": ["vitamin d", "25-hydroxy vitamin d", "25-oh vitamin d", "vit d"],
        "ferritin": ["ferritin, serum"],
        "tibc transferrin binding capacity": ["tibc", "total iron binding capacity"],
        "tsat transferrin saturation": ["transferrin saturation", "tsat", "iron saturation"],
        
        # New Mappings
        "total protein": ["protein, total", "protein total"],
        "ast": ["ast (sgot)", "ast sgot", "sgot", "ast/sgot"],
        "alt": ["alt (sgpt)", "alt sgpt", "sgpt", "alt/sgpt"],
        "total cholesterol tc": ["cholesterol, total", "cholesterol total", "total cholesterol"],
        "hdl": ["hdl cholesterol", "hdl-cholesterol", "hdl cholesterol, direct"],
        "ldl": ["ldl chol calc (nih)", "ldl chol calc nih", "ldl-cholesterol", "ldl cholesterol", "ldl chol calc"],
        "tc hdl": ["t. chol/hdl ratio", "t. chol / hdl ratio", "chol/hdl ratio", "t.chol/hdl", "t.chol/hdl ratio", "cholesterol/hdl ratio", "t. chol/hdl", "t. chol / hdl", "tc/hdl", "tc/hdl ratio"],
        "lymphocytes": ["lymphocytes", "lymphs", "lymphocytes %"],
        "lymphs absolute": ["lymphocytes, absolute", "lymphocytes absolute", "lymphs absolute", "lymphs abs", "lymphocytes abs"],
        "eosinophils": ["eosinophils", "eos", "eosinophils %"],
        "eos absolute": ["eosinophils, absolute", "eosinophils absolute", "eos absolute", "eos abs", "eosinophils abs"],
        "basophils": ["basophils", "basos", "basophils %"],
        "baso absolute": ["basophils, absolute", "basophils absolute", "baso absolute", "baso abs", "basos absolute", "basophils abs", "basos abs"],
        "neutrophils absolute": ["neutrophils, absolute", "neutrophils absolute", "neutrophils abs"],
        "monocytes absolute": ["monocytes, absolute", "monocytes absolute", "monocytes abs", "monocytes(absolute)", "monocytes (abs)", "monocytes(abs)"],
        "immature granulocytes": ["immature grans abs", "immature grans", "immature grans (abs)", "immature granulocytes abs", "immature grans(abs)", "immature granulocytes(abs)"]
    }
    
    if sc in mappings and any(clean_name(alias) == pc for alias in mappings[sc]):
        return True
    if pc in mappings and any(clean_name(alias) == sc for alias in mappings[pc]):
        return True
        
    # Standard prefix/substring matches
    if len(sc) > 3 and len(pc) > 3:
        if sc in pc or pc in sc:
            return True
            
    return False

def format_header_date(date_str):
    """
    Converts YYYY-MM-DD HH:MM AM/PM into:
    Line 1: Year (e.g. 2023)
    Line 2: MonthDay (e.g. Mar21)
    Line 3: HourAM/PM (e.g. 10am)
    """
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d %I:%M %p")
    except ValueError:
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
        except ValueError:
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                dt = datetime.now()
                
    line1 = dt.strftime("%Y")
    line2 = dt.strftime("%b%d")
    hour = dt.strftime("%I").lstrip("0")
    ampm = dt.strftime("%p").lower()
    line3 = f"{hour}{ampm}"
    return f"{line1}\n{line2}\n{line3}"

def create_template(excel_path: str, gender: str = "Female") -> None:
    """
    Creates a new template spreadsheet populated with baseline columns and panels.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blood Test Tracking"
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"
    
    # 1. Base Headers
    ws.append(["TestName", "Normal\nReference Ranges", "Optimal\nRanges"])
    
    # Format Headers
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
    
    # Set Row 1 Height
    ws.row_dimensions[1].height = 40
    
    # Gather panels
    panels = COMMON_PANELS.copy()
    if gender.lower() == "male":
        panels.append(MALE_HORMONE_PANEL)
    else:
        panels.append(FEMALE_HORMONE_PANEL)
        
    # Write panels
    curr_row = 2
    for panel in panels:
        # Write subsection row separator
        ws.cell(row=curr_row, column=1, value=panel["name"])
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=3)
        
        # Style separator row across A-C
        for col in range(1, 4):
            cell = ws.cell(row=curr_row, column=col)
            cell.fill = PANEL_FILL
            cell.font = PANEL_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = CELL_BORDER
        ws.row_dimensions[curr_row].height = 24
        curr_row += 1
        
        # Write tests
        for test_name, normal, optimal in panel["tests"]:
            ws.cell(row=curr_row, column=1, value=test_name)
            ws.cell(row=curr_row, column=2, value=normal)
            ws.cell(row=curr_row, column=3, value=optimal)
            
            # Style cells
            for col in range(1, 4):
                cell = ws.cell(row=curr_row, column=col)
                cell.font = STANDARD_FONT
                # If there are line breaks in normal/optimal ranges, enable wrap text
                cell.alignment = Alignment(
                    horizontal="left" if col == 1 or "\n" in str(cell.value) else "center",
                    vertical="center",
                    wrap_text=True if "\n" in str(cell.value) or col > 1 else False
                )
                cell.border = CELL_BORDER
            
            # Auto height if multiline
            lines = max(str(normal).count("\n"), str(optimal).count("\n"))
            ws.row_dimensions[curr_row].height = max(18, (lines + 1) * 15)
            curr_row += 1
            
    # Auto-fit columns
    autofit_columns(ws)
    
    # Save spreadsheet
    os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)
    wb.save(excel_path)
    wb.close()

def is_cell_merged(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.row >= merged_range.min_row and cell.row <= merged_range.max_row:
            if cell.column >= merged_range.min_col and cell.column <= merged_range.max_col:
                if merged_range.max_col - merged_range.min_col > 0:
                    return True
    return False

def autofit_columns(ws):
    """
    Resizes columns dynamically to fit content nicely.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if is_cell_merged(ws, cell):
                continue
                
            val = str(cell.value or '')
            # Handle line splits
            lines = val.split("\n")
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
                    
        # Apply padding and set column width
        # If it is column B or C, restrict the width to 18 (about 2 inches)
        if col_letter in ["B", "C"]:
            ws.column_dimensions[col_letter].width = 18
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def cleanup_miscellaneous_duplicates(ws):
    """
    Scans the worksheet for any duplicate rows in the 'Miscellaneous / Unmapped Markers'
    section that actually map to standard tests in the panels above.
    Merges their values into the standard rows (if the standard rows don't already have values)
    and deletes the duplicate rows.
    """
    # 1. Find the separator row for Miscellaneous / Unmapped Markers
    misc_separator_row = None
    for row in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        val = cell_a.value
        if val and any(kw in str(val).lower() for kw in ["miscellaneous", "unmapped"]):
            misc_separator_row = row
            break
            
    if not misc_separator_row:
        return
        
    # 2. Gather all standard rows (rows above the separator)
    # We map clean_name(standard_test_name) -> row_index
    standard_rows = {}
    for row in range(2, misc_separator_row):
        cell_a = ws.cell(row=row, column=1)
        val = cell_a.value
        # Skip merged panel headers
        if not val or cell_a.coordinate in ws.merged_cells or is_cell_merged(ws, cell_a):
            continue
        standard_rows[clean_name(val)] = row

    # 3. Scan rows below the separator
    rows_to_delete = []
    # We scan from misc_separator_row + 1 to ws.max_row
    for row in range(misc_separator_row + 1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        val_misc = cell_a.value
        if not val_misc:
            continue
            
        # Check if it matches any standard row
        matched_std_row = None
        for std_clean, std_row_idx in standard_rows.items():
            std_name = ws.cell(row=std_row_idx, column=1).value
            if match_test_name(std_name, val_misc):
                matched_std_row = std_row_idx
                break
                
        if matched_std_row:
            # We found a duplicate! Let's merge values from the miscellaneous row to the standard row
            # For each data column (from column 4 onwards)
            for col in range(4, ws.max_column + 1):
                misc_val = ws.cell(row=row, column=col).value
                std_cell = ws.cell(row=matched_std_row, column=col)
                # If std_cell doesn't have a value but misc_cell does, copy it
                if std_cell.value is None or std_cell.value == "":
                    if misc_val is not None and misc_val != "":
                        std_cell.value = misc_val
                        std_cell.font = BOLD_FONT
                        std_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            rows_to_delete.append(row)
            
    # 4. Delete the duplicate rows in reverse order
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)
        
    # 5. Check if there are any rows remaining under the separator
    new_separator_row = None
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and any(kw in str(val).lower() for kw in ["miscellaneous", "unmapped"]):
            new_separator_row = row
            break
            
    if new_separator_row and new_separator_row == ws.max_row:
        # No rows below the separator, so we delete the separator row itself
        ws.delete_rows(new_separator_row)

def append_data_column(excel_path: str, parsed_data: dict, gender: str = "Female") -> None:
    """
    Appends a new data column on the far right of the spreadsheet.
    """
    if not os.path.exists(excel_path):
        create_template(excel_path, gender)
        
    try:
        wb = openpyxl.load_workbook(excel_path)
    except PermissionError:
        raise PermissionError(
            f"The Excel file '{os.path.basename(excel_path)}' is currently open in Microsoft Excel. "
            "Please close Microsoft Excel and try again."
        )
    ws = wb.active
    ws.freeze_panes = "A2"
    
    # Clean up any existing duplicate rows in the Miscellaneous section
    cleanup_miscellaneous_duplicates(ws)
    
    test_date = parsed_data.get("test_date") or datetime.now().strftime("%Y-%m-%d 08:00 AM")
    raw_results = parsed_data.get("results") or []
    
    # Filter out ignored tests like CHD Risk
    ignored_keywords = ["chd risk", "chdrisk", "estimated chd"]
    results = []
    for res in raw_results:
        p_name = res.get("test_name")
        if p_name and any(kw in clean_name(p_name) for kw in ignored_keywords):
            continue
        results.append(res)
    
    # Find next column index
    new_col_idx = ws.max_column + 1
    new_col_letter = get_column_letter(new_col_idx)
    
    # Write header
    header_val = format_header_date(test_date)
    header_cell = ws.cell(row=1, column=new_col_idx, value=header_val)
    header_cell.fill = HEADER_FILL
    header_cell.font = HEADER_FONT
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_cell.border = CELL_BORDER
    
    # Keep track of matched results so we can append leftovers if necessary
    matched_results = set()
    
    # Fill in matching test cells
    for row in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        test_name_in_sheet = cell_a.value
        
        # Skip merged panel separators
        if cell_a.coordinate in ws.merged_cells or not test_name_in_sheet or is_cell_merged(ws, cell_a):
            # We must style the empty cells in separator rows
            empty_cell = ws.cell(row=row, column=new_col_idx)
            empty_cell.fill = PANEL_FILL
            empty_cell.border = CELL_BORDER
            continue
            
        # Search for matching parsed test result
        matched_val = ""
        for i, res in enumerate(results):
            parsed_name = res.get("test_name")
            parsed_val = res.get("value")
            
            if match_test_name(test_name_in_sheet, parsed_name):
                # Prefer absolute counts or non-percentage values for standard absolute rows
                if not matched_val:
                    matched_val = parsed_val
                else:
                    sc_clean = clean_name(test_name_in_sheet)
                    pc_clean = clean_name(parsed_name)
                    is_std_abs = ("absolute" in sc_clean or "abs" in sc_clean or "granulocytes" in sc_clean)
                    is_new_abs = ("abs" in pc_clean or "absolute" in pc_clean)
                    if is_std_abs and is_new_abs:
                        matched_val = parsed_val
                matched_results.add(i)
                
        # Write value
        val_cell = ws.cell(row=row, column=new_col_idx, value=matched_val)
        val_cell.font = BOLD_FONT
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = CELL_BORDER
        
    # Optional: If there are unmatched parsed results, append them at the end of the sheet
    # under a "Unmapped Markers" panel. Let's do this to make sure we don't lose any data!
    unmatched_indices = set(range(len(results))) - matched_results
    if unmatched_indices:
        # Check if we already have an "Unmapped / Miscellaneous Markers" section
        unmapped_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and any(kw in str(val).lower() for kw in ["miscellaneous", "unmapped"]):
                unmapped_row = row
                break
                
        if not unmapped_row:
            # Create a separator
            ws.append(["Miscellaneous / Unmapped Markers"])
            unmapped_row = ws.max_row
            ws.merge_cells(start_row=unmapped_row, start_column=1, end_row=unmapped_row, end_column=new_col_idx)
            for col in range(1, new_col_idx + 1):
                cell = ws.cell(row=unmapped_row, column=col)
                cell.fill = PANEL_FILL
                cell.font = PANEL_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = CELL_BORDER
            ws.row_dimensions[unmapped_row].height = 24
            
        for i in unmatched_indices:
            res = results[i]
            p_name = res.get("test_name")
            p_val = res.get("value")
            
            # Append new row
            new_row_idx = ws.max_row + 1
            ws.cell(row=new_row_idx, column=1, value=p_name)
            ws.cell(row=new_row_idx, column=2, value="")
            ws.cell(row=new_row_idx, column=3, value="")
            
            # Style baseline columns in this new row
            for col in range(1, 4):
                cell = ws.cell(row=new_row_idx, column=col)
                cell.font = STANDARD_FONT
                cell.border = CELL_BORDER
                cell.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
                
            # Write data column
            val_cell = ws.cell(row=new_row_idx, column=new_col_idx, value=p_val)
            val_cell.font = BOLD_FONT
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            val_cell.border = CELL_BORDER
            
            # Apply border to all other empty columns in this new row
            for col in range(4, new_col_idx):
                ws.cell(row=new_row_idx, column=col).border = CELL_BORDER
                
            ws.row_dimensions[new_row_idx].height = 18
            
    # Calculate ratios dynamically if they are missing but constituent values exist
    row_tc = None
    row_hdl = None
    row_tg = None
    row_ldl = None
    
    row_tc_hdl = None
    row_tg_hdl = None
    row_ldl_hdl = None
    
    for row in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        val_a = cell_a.value
        if not val_a or cell_a.coordinate in ws.merged_cells or is_cell_merged(ws, cell_a):
            continue
        
        name = clean_name(val_a)
        if name == "total cholesterol tc":
            row_tc = row
        elif name == "hdl":
            row_hdl = row
        elif name == "triglycerides tg":
            row_tg = row
        elif name == "ldl":
            row_ldl = row
        elif name == "tc hdl":
            row_tc_hdl = row
        elif name == "tg hdl":
            row_tg_hdl = row
        elif name == "ldl hdl":
            row_ldl_hdl = row
            
    def get_float_val(row_idx):
        if not row_idx:
            return None
        cell_val = ws.cell(row=row_idx, column=new_col_idx).value
        if not cell_val:
            return None
        # Try to parse float
        s = str(cell_val).strip()
        cleaned = []
        for c in s:
            if c.isdigit() or c == '.':
                cleaned.append(c)
            else:
                break
        try:
            return float("".join(cleaned))
        except ValueError:
            return None

    val_tc = get_float_val(row_tc)
    val_hdl = get_float_val(row_hdl)
    val_tg = get_float_val(row_tg)
    val_ldl = get_float_val(row_ldl)
    
    if val_hdl and val_hdl > 0:
        # TC/HDL Ratio
        if row_tc_hdl:
            if val_tc:
                ratio = round(val_tc / val_hdl, 2)
                ws.cell(row=row_tc_hdl, column=new_col_idx, value=ratio).font = BOLD_FONT
                ws.cell(row=row_tc_hdl, column=new_col_idx).alignment = Alignment(horizontal="center", vertical="center")
        # TG/HDL Ratio
        if row_tg_hdl:
            if val_tg:
                ratio = round(val_tg / val_hdl, 2)
                ws.cell(row=row_tg_hdl, column=new_col_idx, value=ratio).font = BOLD_FONT
                ws.cell(row=row_tg_hdl, column=new_col_idx).alignment = Alignment(horizontal="center", vertical="center")
        # LDL/HDL Ratio
        if row_ldl_hdl:
            if val_ldl:
                ratio = round(val_ldl / val_hdl, 2)
                ws.cell(row=row_ldl_hdl, column=new_col_idx, value=ratio).font = BOLD_FONT
                ws.cell(row=row_ldl_hdl, column=new_col_idx).alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    autofit_columns(ws)
    
    try:
        wb.save(excel_path)
    except PermissionError:
        raise PermissionError(
            f"The Excel file '{os.path.basename(excel_path)}' is currently open in Microsoft Excel. "
            "Please close Microsoft Excel and try again."
        )
    wb.close()

if __name__ == "__main__":
    # Test block
    test_path = r"c:\Users\admin\Desktop\emma-ai-capstone\test_labs_female.xlsx"
    if os.path.exists(test_path):
        os.remove(test_path)
        
    print("Creating female template...")
    create_template(test_path, "Female")
    
    # Mock data to append
    mock_data = {
        "test_date": "2023-03-21 09:15 AM",
        "results": [
            {"test_name": "WBC", "value": "5.2"},
            {"test_name": "RBC", "value": "4.15"},
            {"test_name": "Hemoglobin", "value": "13.2"},
            {"test_name": "Estradiol (E2)", "value": "115.4"},
            {"test_name": "Progesterone", "value": "18.2"},
            {"test_name": "Glucose", "value": "78"},
            {"test_name": "Some Random Test", "value": "9.9"} # Should go to Miscellaneous
        ]
    }
    
    print("Appending mock data...")
    append_data_column(test_path, mock_data, "Female")
    print(f"Test completed. Sheet saved at {test_path}")
